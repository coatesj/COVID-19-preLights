# Script that looks through Covid19 preprints and checks whether
# they have been peer-reviewed and published in a journal
# Author: Lars Hubatsch, hubatsch@gmail.com

from bs4 import BeautifulSoup
from numpy import savetxt
from openpyxl import load_workbook
from selenium import webdriver

# Here goes the path to your excel file
wb = load_workbook('/Users/hubatsch/Desktop/Covid19.xlsx')
ws = wb['Sheet1']

# Loop through all rows in excel file, get link from first column
# check whether page contains 'Now published in', append number
# to published depending on outcome:
#  0 ... no link in row
#  1 ... link in row with peer-reviewed version
# -1 ... link in row without peer-reviewed version
published = []
for row in range(1, ws.max_row+1):
    try:
        print(ws.cell(row=row, column=1).hyperlink.target)
        base_url = ws.cell(row=row, column=1).hyperlink.target
        driver = webdriver.PhantomJS()
        driver.get(base_url)
        driver.implicitly_wait(100)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        a = soup.find_all('div', attrs={'class': 'pub_jnl'})
        print(str(a[0]).find('Now published in'))
        if str(a[0]).find('Now published in') == -1:
            published.append(-1)
        else:
            published.append(1)
    except (ValueError, AttributeError):
        print('No link in row ' + str(row))
        published.append(0)
savetxt('published.txt', published, fmt='%i',)